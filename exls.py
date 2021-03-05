import openpyxl
import subprocess
import re


def get_value_excel(filename, x, y):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['user']
    wb.close()
    return Sheet1.cell(row=x, column=y).value

def update_value_excel(filename, x, y, value):
    wb = openpyxl.load_workbook(filename)
    Sheet1 = wb['user']
    Sheet1.cell(row=x, column=y).value = value
    wb.close()
    wb.save(filename)
filename = "user.xlsx"
last_row= int(input('Number column is: '))
r1=last_row +1
for x in range(3,r1):
    cno= get_value_excel(filename,x,1)
    cid= get_value_excel(filename,x,2)
    cname = get_value_excel(filename,x,3)
    cdepartment= get_value_excel(filename,x,4)
    cemail = get_value_excel(filename,x, 5)
    print('no:',cno)
    a1 = f'dsadd user "cn={cid},ou=hansol,dc=hansol,dc=local" -fn {cid} -display "{cname}" -pwd Abc13579 -office {cdepartment} -email "{cemail}" -mustchpwd yes'

    process = subprocess.Popen(f"{a1}",
                           stdout=subprocess.PIPE,
                           stderr=subprocess.PIPE,
                           shell=True
                           )
    while process.poll() is None:
        lineerr = process.stderr.read()        
        lineout = process.stdout.read()   
        print(lineerr.strip().decode('utf-8'))
        print(lineout.strip().decode('utf-8'))
        update_value_excel(filename,x,6, xz)

    
    #input('pause')
