from openpyxl.descriptors.base import Length
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import requests
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import re

filename = 'listprinter - Copy (2).xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb.active
for y in range(1,sheet.max_column):
    max_length = 0
    column = sheet.cell(row=3, column=y).column
    print(column)
    for cell in y:
        if cell.coordinate in sheet.merged_cells: # not check merge_cells
            continue
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    print(column)
    sheet.column_dimensions[column].width = adjusted_width
wb.save(filename)

