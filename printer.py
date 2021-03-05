from openpyxl.descriptors.base import Length
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import requests
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import re


def cell_coordinate(filename, x, y):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]
    cell =sheet.cell(row=x, column=y)
    wb.close()
    return cell.coordinate

def get_value_excel(filename, x, y):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]
    wb.close()
    return sheet.cell(row=x, column=y).value

def last_row_col(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]
    row_count = sheet.max_row
    column_count = sheet.max_column
    wb.close()
    return (row_count, column_count)

def update_value_excel(filename, x, y, value):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]
    sheet.cell(row=x, column=y).value = value    
    wb.close()
    wb.save(filename)   

def set_border(filename, x, y):
    border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]
    sheet.cell(row=x, column=y).border =border
    wb.close()
    wb.save(filename)
def fit_col (filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.worksheets[0]
    column_widths = []
    for row in range (1,1,sheet.max_row):
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]

    for i, column_width in enumerate(column_widths):
        sheet.column_dimensions[get_column_letter(i+1)].width = column_width
def getinfo(ip):
    ipp=f'http://{ip}'
    try:
        request = requests.get(ipp) #Here is where im getting the error
        if request.status_code == 200:            
            driver.get(ipp)
            ele = driver.find_element_by_id('LogBox')
            ele.send_keys('initpass')
            ele1 = driver.find_element_by_id('login')
            ele1.click()
            ipmi = f'http://{ip}/general/information.html?kind=item'
            driver.get(ipmi)              
            page_counter =driver.find_element_by_xpath('//*[@id="pageContents"]/form/div[6]/dl/dd[1]')
            pc= page_counter.text    
            return (pc)            
    except:
        return ('Connection error')

    
'''
ip = "172.26.5.222"
page_counter = getinfo(ip)
print(page_counter)
'''
filename = 'listprinter.xlsx'
last_row, last_col = last_row_col(filename)
tpm = "Total Pages/Month"
pc= "Page Counter"
update_value_excel(filename, 3, last_col+2, tpm)
update_value_excel(filename, 3, last_col+1, pc)

#hide chrome
options = webdriver.ChromeOptions()
options.add_argument('headless')
options.add_argument('window-size=1920x1080')
options.add_argument("disable-gpu")
# OR options.add_argument("--disable-gpu")
driver = webdriver.Chrome('chromedriver.exe', chrome_options=options)
for x in range (4,last_row +1):
    ip = get_value_excel(filename, x, 4)
    page_counter = getinfo(ip)
    st= cell_coordinate(filename,x, last_col+1)
    sbt= cell_coordinate(filename,x, last_col-1) 
    ct = f"={st}-{sbt}"
    update_value_excel(filename, x, last_col+2, ct)
    set_border(filename, x, last_col+2)
    update_value_excel(filename, x, last_col+1, page_counter)
    set_border(filename, x, last_col+1)
       
driver.quit()
fit_col(filename)